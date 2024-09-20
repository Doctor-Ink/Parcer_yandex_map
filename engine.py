import os
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException


def get_name_company(driver, logger, card):
    name_company = ''
    try:
        name_company = driver.find_element(By.CLASS_NAME, 'card-title-view__title-link').text.strip()
        logger.info("Name company is found for the card: %s", name_company)
    except NoSuchElementException:
        logger.warning("Name company not found for the card: %s", card)
    except Exception as exc:
        logger.error("An unexpected error occurred while trying to get the name compamy for the card: %s %s", card, exc)
    return name_company


def get_addresses(driver, logger, card):
    # забираем адрес компании
    address = ''
    try:
        address = driver.find_element(By.CLASS_NAME, 'business-contacts-view__address-link').text.strip()
        logger.info("Address found for the card: %s", address)
    except NoSuchElementException:
        logger.warning("Address not found for the card: %s", card)
    except Exception as exc:
        logger.error("An unexpected error occurred while trying to get the address for the card: %s %s", card, exc)
    return address


def get_contacts(driver, logger, card):
    wait = WebDriverWait(driver, 1)
    # забираем телефонные номера
    contacts = ''
    try:
        # ожидаем пока появится кнопка "показать телефон"
        show_phone = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'card-phones-view__more')))
        show_phone.click()

        logger.info("The button - Показать номер PUSH ON - %s", card)
        arrow_down = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//div[@class='card-phones-view']//div[@class='card-feature-view__additional']"))
        )

        logger.info("Нажимаем стрелку вниз для телефонов - %s", card)
        arrow_down.click()

        logger.info("Нажали стрелку вниз для телефонов - %s", card)
        # ожидаем когда появится стрелка вверх - для полной загрузки контактов
        wait.until(EC.presence_of_element_located(
            (By.CLASS_NAME, "card-feature-view__arrow _view_up"))
        )
        logger.info("Дождались появления стрелки вверх КОНТАКТЫ - %s", card)
    except TimeoutException:
        logger.info("No 'Show phone' button not found for card - %s", card)

    try:
        info_phones = driver.find_elements(
            By.XPATH, "//div[@class='card-phones-view']//div[@class='card-feature-view__content']"
        )
        for phone_number in info_phones:
            contacts += phone_number.text.strip() + ' '
        logger.info("Phone numbers is done!!! for card - %s", card)
    except Exception as exc:
        logger.warning("Phone numbers company is empty - %s  %s", card, exc)
    logger.info("Contacts found for the card: %s", contacts)
    return contacts


def get_site(driver, logger, card):
    # забираем сайт компании
    site = ''
    try:
        site = driver.find_element(By.CLASS_NAME, 'business-urls-view__text').text.strip()
        logger.info("Site found for the card: %s", site)
    except NoSuchElementException:
        logger.warning("Site not found for the card: %s", card)
    except Exception as exc:
        logger.error("An unexpected error occurred while trying to get the address for the card: %s %s", card, exc)
    return site


def get_soc_network(driver, logger, card):
    soc_network = ''
    # забираем ссылки на соцсети
    try:
        social_medias = driver.find_elements(By.CLASS_NAME, 'business-contacts-view__social-button')
        if social_medias:
            for elem in social_medias:
                soc_network += elem.find_element(By.TAG_NAME, 'a').get_attribute('href') + ' '
        logger.info("Social network found for the card: %s", soc_network)
    except NoSuchElementException:
        logger.warning("Social network not found for the card: %s", card)
    except Exception as exc:
        logger.error("An unexpected error occurred while trying to get the Social network for the card: %s %s", card, exc)

    return soc_network


def get_coordinates(driver, logger, card):
    coordinates = ''
    # # Забираем координаты
    try:
        my_url = driver.current_url
        # https://yandex.by/maps/org/banya_na_kurskoy/232229006327/?ll=35.432039%2C52.331216&z=9
        some_list = my_url.split('/')[-1].split('.')
        coordinates = some_list[0][-2:] + '.' + some_list[1][:6] + ', ' + some_list[1][-2:] + '.' + some_list[2][:6]
        logger.info("coordinates found for the card: %s", coordinates)
    except Exception as exc:
        logger.warning("Coordinates is EMPTY - %s ", card)
    return coordinates


def cur_url(driver, logger, card):
    url = ''
    # забираем урл на карточку
    try:
        url = driver.find_element(By.CLASS_NAME, 'card-title-view__title-link').get_attribute('href')
    except NoSuchElementException:
        logger.warning("Current url not found for the card: %s", card)
    except Exception as exc:
        logger.error("An unexpected error occurred while trying to get the CURREN URL for the card: %s %s", card, exc)
    return url


def time_track(func):
    # функция-декаратор, которая считает время работы
    def surogate(*args, **kwargs):
        start_time = time.time()

        result = func(*args, **kwargs)

        end_time= time.time()
        result_time = end_time - start_time
        print(f'Скрипт отработал - {round(result_time, 2)} секунды')
        return result
    return surogate


def create_append_df_to_excel(filename, df, sheet_name='sheet1'):
    # функция принимем датафрэйм пандас и записывает в эксель
    # Проверка существования файла
    if not os.path.exists(filename):
        df.to_excel(filename, sheet_name=sheet_name, index=False)
    # добавление в эксель файл
    else:
        book = load_workbook(filename)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            sheet = book.create_sheet(sheet_name)

        # Append the DataFrame rows to the sheet
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        # Save the workbook
        book.save(filename)